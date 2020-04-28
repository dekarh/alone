from collections import OrderedDict
import os

from datetime import datetime, timedelta, time, date
import openpyxl

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidgetItem

from mysql.connector import MySQLConnection

from alone_win import Ui_Form

from lib import read_config, l, s, fine_phone, format_phone, fine_snils

class MainWindowSlots(Ui_Form):   # Определяем функции, которые будем вызывать в слотах

    def setupUi(self, form):
        Ui_Form.setupUi(self,form)
        self.client_id = None
        self.hasFileFolder = False
        self.dbconfig_crm = read_config(filename='alone.ini', section='crm')
        self.dbconfig_alone = read_config(filename='alone.ini', section='alone')
        # Соптимизировали неиспользующийся огромный словарь
        q2 = """ 
        self.alone_files = {}
        with open("all_files.txt", "rt") as file_all:
            for i, line in enumerate(file_all):
                if i > 1:
                    if len(line.split('/')) > 2 and line.find('search') == -1:
                        file_name = line.split('.wav')[0].split('/')[2].lower()
                        path_name = line.split('./recup_dir.')[1].split('/')[0].replace('/n','')
                        if self.alone_files.get(file_name, None):
                            self.alone_files[file_name].append(path_name)
                        else:
                            self.alone_files[file_name] = [path_name]
        """
        wb = openpyxl.load_workbook(filename='нужноАудио.xlsx', read_only=True)
        ws = wb[wb.sheetnames[0]]
        self.not_finded_snilses = []
        for i, row in enumerate(ws):
            if i:
                for j, cell in enumerate(row):
                    if j == 1:
                        self.not_finded_snilses.append(l(cell.value))
                        break
        self.twRezkeyPressEventMain = self.twRez.keyPressEvent
        self.twRez.keyPressEvent = self.twRezkeyPressEvent
        self.clbSave.setEnabled(False)
        self.contracts = {None:None}
        self.clbReport2xlsx.setEnabled(False)
        self.threads = []
        self.progressBar.hide()
        return

    def click_clbLoadBLUE(self):
        # С галочкой для выгрузки внутри СатурнОПС
        with_exch_tuple = (11935183763, 4860950377, 12423639535, 4378543274, 11674757986, 2950346044, 8978383958,
                            12329988078, 12484652164, 14122520601, 4409787070, 11282485644, 6433441240, 11436693958,
                            6383908891, 7523343148, 7210778646, 8008494357, 7395473199, 2855105242, 14210167703,
                            6845027476, 5946598220, 579266261, 15682898527, 13807927887, 14604830851, 15193939996,
                            5564802672, 2482394149, 10585759885, 5328385664, 5332154520, 2707298757, 5703889279,
                            5714645464, 6929973837, 12545430336, 5637917996, 13069208643, 395657254, 11725013617,
                            4713439549, 14257079462, 13144576844, 7692235492, 6730208543, 10634600820, 6270537248,
                            14299059086, 5475460775, 14115615722, 9962245110, 10883327261, 12367816669, 5644764281,
                            5535908271, 15313860747, 11811393735, 12303112072, 6954910100, 10508187732, 2439134936,
                            6300755728, 4171238826, 12891654086, 1007496199, 13294926074, 1007484091, 7698642636,
                            3955799821, 12460657041, 7013869246, 7191985798, 13164142318, 3987760923, 11631685342,
                            13721572750, 7668297829, 6093750059, 11843929576, 10334126898, 7244572664, 4113831911,
                            14512255731, 11911348434, 6566976117, 5637841888, 5310604503, 7901302948, 11044685624,
                            7962810099, 13831199262, 15049928883, 9664279924, 7332090233, 13826618073, 13197913578,
                            11393718659, 13313258619, 1497799804, 10882201739, 9576905023, 16687370215, 4375511549,
                            1100125843, 12858423583, 13665396995, 14565028664, 7003158813, 6938659832, 4783949614,
                            12533646141, 5235408533, 3730548442, 11439507243, 2872166158, 4806851167, 13625814559,
                            4590414249, 13840814964, 3439780470, 10878243375, 14659012468, 11426599859, 9420992482,
                            5937237086, 4928277493, 11547334342, 11756026043, 13600336613, 7475331477, 12978031989,
                            4311049094, 8053356552, 4031186802, 5670609368, 938183458, 13526978890, 13845084978,
                            10061726096, 2576587383, 7970843312, 3316903022, 15675289309, 4463668775, 4764345575,
                            3353497551, 15435683477, 1225701493, 4183723546, 9774599551, 13442150417, 14025523817,
                            13146129021, 13142917228, 13220648919, 5091695261, 12582210937, 14835250163, 8731279789,
                            11266925452, 13524971967, 12737228561, 6221687747, 7124088438, 7061077435, 13648873299,
                            6806384279, 13305312199, 11942497474, 2722899770, 14586901394, 7344588080, 11858377695,
                            4818019155, 11331668119, 6010125276)
        with_exch_dict = {11935183763: 859261, 4860950377: 869902, 12423639535: 871976, 4378543274: 872457,
                          11674757986: 872644, 2950346044: 872666, 8978383958: 873774, 12329988078: 873822,
                          12484652164: 874171, 14122520601: 881983, 4409787070: 977335, 11282485644: 979296,
                          6433441240: 980300, 11436693958: 1119772, 6383908891: 1301793, 7523343148: 1590961,
                          7210778646: 1597286, 8008494357: 1610109, 7395473199: 1613319, 2855105242: 1613736,
                          14210167703: 1626074, 6845027476: 1655059, 5946598220: 1674042, 579266261: 1723842,
                          15682898527: 1741070, 13807927887: 1763616, 14604830851: 1763729, 15193939996: 1764130,
                          5564802672: 1764968, 2482394149: 1765638, 10585759885: 1782249, 5328385664: 1793727,
                          5332154520: 1800748, 2707298757: 1800761, 5703889279: 1801175, 5714645464: 1802907,
                          6929973837: 1804308, 12545430336: 1816436, 5637917996: 1837165, 13069208643: 1837284,
                          395657254: 1904617, 11725013617: 1930269, 4713439549: 1930292, 14257079462: 1944022,
                          13144576844: 1955358, 7692235492: 1957139, 6730208543: 1959634, 10634600820: 1959707,
                          6270537248: 1959735, 14299059086: 1972354, 5475460775: 1975658, 14115615722: 1975703,
                          9962245110: 1980978, 10883327261: 1989387, 12367816669: 1989451, 5644764281: 1989473,
                          5535908271: 1990840, 15313860747: 1992111, 11811393735: 1993646, 12303112072: 1993744,
                          6954910100: 1993772, 10508187732: 2006890, 2439134936: 2008888, 6300755728: 2013115,
                          4171238826: 2013279, 12891654086: 2015375, 1007496199: 2016158, 13294926074: 2017013,
                          1007484091: 2020665, 7698642636: 2021140, 3955799821: 2023333, 12460657041: 2024811,
                          7013869246: 2025072, 7191985798: 2026299, 13164142318: 2026924, 3987760923: 2027352,
                          11631685342: 2028836, 13721572750: 2031319, 7668297829: 2031928, 6093750059: 2031933,
                          11843929576: 2031946, 10334126898: 2032271, 7244572664: 2032944, 4113831911: 2033629,
                          14512255731: 2034049, 11911348434: 2034320, 6566976117: 2034781, 5637841888: 2035149,
                          5310604503: 2035594, 7901302948: 2035641, 11044685624: 2036311, 7962810099: 2038071,
                          13831199262: 2039566, 15049928883: 2041736, 9664279924: 2053395, 7332090233: 2054090,
                          13826618073: 2054821, 13197913578: 2054943, 11393718659: 2055269, 13313258619: 2058049,
                          1497799804: 2061146, 10882201739: 2061982, 9576905023: 2065120, 16687370215: 2066221,
                          4375511549: 2069074, 1100125843: 2078167, 12858423583: 2079660, 13665396995: 2085392,
                          14565028664: 2085818, 7003158813: 2085853, 6938659832: 2092425, 4783949614: 2094748,
                          12533646141: 2108113, 5235408533: 2110589, 3730548442: 2113161, 11439507243: 2116927,
                          2872166158: 2120147, 4806851167: 2126661, 13625814559: 2129329, 4590414249: 2212436,
                          13840814964: 2257365, 3439780470: 2257811, 10878243375: 2271981, 14659012468: 2716285,
                          11426599859: 2775404, 9420992482: 2786723, 5937237086: 2973878, 4928277493: 3028453,
                          11547334342: 3028638, 11756026043: 3028675, 13600336613: 3028912, 7475331477: 3030523,
                          12978031989: 3033191, 4311049094: 3035576, 8053356552: 3035890, 4031186802: 3037823,
                          5670609368: 3038976, 938183458: 3042515, 13526978890: 3043352, 13845084978: 3044522,
                          10061726096: 3046847, 2576587383: 3047386, 7970843312: 3048632, 3316903022: 3049796,
                          15675289309: 3052953, 4463668775: 3054397, 4764345575: 3054693, 3353497551: 3054833,
                          15435683477: 3054981, 1225701493: 3055017, 4183723546: 3055589, 9774599551: 3056500,
                          13442150417: 3056930, 14025523817: 3057741, 13146129021: 3057858, 13142917228: 3058711,
                          13220648919: 3058750, 5091695261: 3060223, 12582210937: 3060280, 14835250163: 3061172,
                          8731279789: 3062427, 11266925452: 3062785, 13524971967: 3063603, 12737228561: 3505873,
                          6221687747: 3555521, 7124088438: 3557806, 7061077435: 3565506, 13648873299: 3571520,
                          6806384279: 3667626, 13305312199: 3700296, 11942497474: 3701792, 2722899770: 3705698,
                          14586901394: 3711556, 7344588080: 3860823, 11858377695: 3861602, 4818019155: 3876326,
                          11331668119: 3923544, 6010125276: 3925346}
        dbconn = MySQLConnection(**self.dbconfig_crm)
        cursor = dbconn.cursor()
        cursor.execute('SELECT callcenter_id, path, file FROM lekarh.alone_remont')
        rows = cursor.fetchall()
        alone_remonts = {}
        for row in rows:
            if alone_remonts.get(row[0], None):
                alone_remonts[row[0]] += [os.path.join('/back/recup_dir.' + str(row[1]), row[2])]
            else:
                alone_remonts[row[0]] = [os.path.join('/back/recup_dir.' + str(row[1]), row[2])]
        cursor = dbconn.cursor()
        cursor.execute('SELECT cl.number, ca.id, concat_ws(" ", cl.p_surname, cl.p_name, cl.p_lastname), '
                       'cl.b_date, cl.p_service_address FROM saturn_crm.callcenter AS ca '
                       'LEFT JOIN saturn_crm.contracts AS co ON ca.contract_id = co.id '
                       'LEFT JOIN saturn_crm.clients AS cl ON co.client_id = cl.client_id '
                       'WHERE cl.subdomain_id = 13 and ca.id < 3933226 and cl.number IN (' +
                       ','.join([str(q) for q in self.not_finded_snilses]) + ') ORDER BY ca.id DESC')
        rows = cursor.fetchall()
        wb_paths = openpyxl.Workbook(write_only=True)
        ws_paths = wb_paths.create_sheet('Прослушивание по папкам')
        ws_paths.append(['№ п/п', 'Файл', 'СНИЛС', 'Ф.И.О.', 'День рождения', 'Прописка'])
        ws_unknowns = wb_paths.create_sheet('Аудиозаписи из неопределившейся части файлопомойки')
        ws_unknowns.append(['callcenter_id', 'СНИЛС', 'Дата звонка', 'Ф.И.О.', 'день рождения', 'Прописка'])
        snilses_writed = ()
        for row in rows:
            if row[0] not in snilses_writed:
                if row[0] in with_exch_tuple:
                    if row[1] == with_exch_dict[row[0]]:
                        if alone_remonts.get(row[1], None):
                            ws_paths.append([])
                            for alone_remont in alone_remonts[row[1]]:
                                ws_paths.append([alone_remont, fine_snils(row[0]), row[2], row[3], row[4]])
                        else:
                            ws_unknowns.append([row[1], fine_snils(row[0]), row[2], row[3], row[4]])
                        snilses_writed += (row[0],)
                else:
                    if alone_remonts.get(row[1], None):
                        ws_paths.append([])
                        for alone_remont in alone_remonts[row[1]]:
                            ws_paths.append([alone_remont, fine_snils(row[0]), row[2], row[3], row[4]])
                    else:
                        ws_unknowns.append([row[1], fine_snils(row[0]), row[2], row[3], row[4]])
                    snilses_writed += (row[0],)
        wb_paths.save('потерянныеАудио-Отчет.xlsx')
        return

    def click_clbNotFindedXLSX(self):
        wb = openpyxl.load_workbook(filename='нужноАудио.xlsx', read_only=True)
        ws = wb[wb.sheetnames[0]]
        dbconn = MySQLConnection(**self.dbconfig_crm)
        cursor = dbconn.cursor()
        self.progressBar.setMaximum(len(self.not_finded_snilses) - 1)
        self.progressBar.show()
        pathDataDate = {}
        # Создаем словарь pathDataDate[path][snils]
        for i, snils in enumerate(self.not_finded_snilses):
            self.progressBar.setValue(i)
            sql = 'SELECT cl.client_id, ca.client_phone, ca.inserted_date, ca.exchangeable, ' \
                  'concat_ws(" ", cl.p_surname, cl.p_name, cl.p_lastname), cl.b_date, cl.p_service_address ' \
                  'FROM saturn_crm.callcenter AS ca ' \
                  'LEFT JOIN saturn_crm.contracts AS co ON ca.contract_id = co.id ' \
                  'LEFT JOIN saturn_crm.clients AS cl ON co.client_id = cl.client_id ' \
                  'WHERE cl.number = %s and cl.subdomain_id = 13'
            cursor.execute(sql, (l(snils),))
            rows = cursor.fetchall()
            if len(rows):
                fio = rows[0][4]
                birthday = rows[0][5]
                address = rows[0][6]
                data = datetime(2001, 1, 1, 0, 0)
                has_checked = False
                for row in rows:
                    if row[2] > data:
                        data = row[2]
                    if row[3]:
                        has_checked = True
                        checked_row = row
                if has_checked:
                    data = checked_row[2]
                finded = False
                for thread in self.threads:
                    if data > thread['start'] and data < thread['end']:
                        finded = True
                        max_date_delta = timedelta(days=10000)
                        min_date_delta = timedelta(days=10000)
                        for path in thread['pathsDates']:
                            for call_date in thread['pathsDates'][path]:
                                if call_date > data and (call_date - data) < max_date_delta:
                                    max_date_delta = call_date - data
                                    max_path = path
                                if call_date < data and (data - call_date) < min_date_delta:
                                    min_date_delta = data - call_date
                                    min_path = path
                        for path in range(min_path, max_path + 1):
                            if pathDataDate.get(path, None):
                                pathDataDate[path][snils] = [path, fine_snils(snils), fio, birthday, address]
                            else:
                                pathDataDate[path] = {snils: [path, fine_snils(snils), fio, birthday, address]}
                if not finded:
                    print('Не найдено', fine_snils(snils), data, fio, birthday, address)
            else:
                self.lbDateTime.setText('Нет такого СНИЛС в БД')
        # Сортируем
        #pathDataDate_sorted = OrderedDict(sorted(pathDataDate.items(), key=lambda t: t[0]))

        # По всем звонкам по СНИЛС из списка ненайденных формируем словарь
        #                                                    not_finded[СНИЛС]={callcenter_id: длительность}
        not_finded = {}
        cursor = dbconn.cursor()
        cursor.execute('SELECT cl.number, ca.id, ca.inserted_date, ca.updated_date FROM saturn_crm.callcenter AS ca '
              'LEFT JOIN saturn_crm.contracts AS co ON ca.contract_id = co.id '
              'LEFT JOIN saturn_crm.clients AS cl ON co.client_id = cl.client_id '
              'WHERE cl.subdomain_id = 13 and cl.number IN (' + ','.join([str(q) for q in self.not_finded_snilses]) +
              ')')
        rows = cursor.fetchall()
        for row in rows:
            if row[3] and row[2]:
                if not_finded.get(row[0], None):
                    not_finded[row[0]][row[1]] = row[3] - row[2]
                else:
                    not_finded[row[0]] = {row[1]: row[3] - row[2]}

        # Перебирая 2,5 млн файлов из файлопомойки определять папка-имя-длительность
        finded4not_finded = []
        self.progressBar.setMaximum(2132814)
        self.progressBar.setValue(0)
        with open('2020-04-26_01-24_mp3wav.csv', 'rt') as file_all:
            for i, line in enumerate(file_all):
                if not i % 1000:
                    self.progressBar.setValue(i)
                if len(line.split('/')) > 2 and line.find('search') == -1:
                    file_name = os.path.basename(line.split('\t')[0])
                    path_name = int(os.path.dirname(line.split('\t')[0]).split('/back/recup_dir.')[1])
                    size = int(line.split('\t')[1])
                    duration = float(line.split('\t')[2])
                    # В соответствии с папкой файлов из файлопомойки проверить все callcenter_id из not_finded[СНИЛС]
                    if pathDataDate.get(path_name, None):
                        for snils in pathDataDate[path_name]:
                            for callcenter_id in not_finded[snils]:
                                if abs(not_finded[snils][callcenter_id] - timedelta(seconds=duration)) < \
                                        timedelta(seconds=0.1):
                                    finded4not_finded.append((path_name, file_name, callcenter_id, snils))
                                    pass
                if not (len(finded4not_finded) % 1000) and len(finded4not_finded):
                    cursor = dbconn.cursor()
                    cursor.executemany('INSERT INTO lekarh.alone_finded (path, file, callcenter_id, snils) '
                                       'VALUES (%s, %s, %s, %s)', finded4not_finded)
                    dbconn.commit()
                    finded4not_finded = []
        self.progressBar.hide()
        return

    def click_clbSNILS(self):
        self.lbDateTime.setText('')
        if l(self.leSNILS.text()) < 10000 or l(self.leSNILS.text()) > 99999999999:
            self.lbDateTime.setText('')
        else:
            dbconn = MySQLConnection(**self.dbconfig_crm)
            cursor = dbconn.cursor()
            sql = 'SELECT cl.client_id, ca.client_phone, ca.inserted_date, ca.exchangeable ' \
                  'FROM saturn_crm.callcenter AS ca ' \
                  'LEFT JOIN saturn_crm.contracts AS co ON ca.contract_id = co.id ' \
                  'LEFT JOIN saturn_crm.clients AS cl ON co.client_id = cl.client_id ' \
                  'WHERE cl.number = %s and cl.subdomain_id = 6'
            cursor.execute(sql, (l(self.leSNILS.text()),))
            rows = cursor.fetchall()
            if len(rows):
                data = datetime(2001,1,1,0,0)
                has_checked = False
                for row in rows:
                    if row[2] > data:
                        data = row[2]
                    if row[3]:
                        has_checked = True
                        checked_row = row
                if has_checked:
                    data = checked_row[2]
                textDateTime = data.strftime('%d.%m.%Y') + ' прослушать папки'
                finded = False
                for thread in self.threads:
                    if data > thread['start'] and data < thread['end']:
                        finded = True
                        max_date_delta = timedelta(days=10000)
                        min_date_delta = timedelta(days=10000)
                        for path in thread['pathsDates']:
                            for call_date in thread['pathsDates'][path]:
                                if call_date > data and (call_date - data) < max_date_delta:
                                    max_date_delta = call_date - data
                                    max_path = path
                                if call_date < data and (data - call_date) < min_date_delta:
                                    min_date_delta = data - call_date
                                    min_path = path
                        if textDateTime[-16:] == 'прослушать папки':
                            textDateTime += ' c ' + str(min_path) + ' по ' + str(max_path)
                        else:
                            textDateTime += ', c ' + str(min_path) + ' по ' + str(max_path)
                if not finded:
                    textDateTime += ' недостаточно информации для определения'
                self.lbDateTime.setText(textDateTime)

            else:
                self.lbDateTime.setText('Нет такого СНИЛС в БД')
            return

    def twRezkeyPressEvent(self,e):
        self.twRezkeyPressEventMain(e)
        if e.key() == Qt.Key_Down or e.key() == Qt.Key_Up:
            self.click_twRez(index=self.twRez.model().index(self.twRez.currentRow(), 0))

    def click_twRez(self, index=None): # Сделать кнопку Сохранить активной если есть файл, папка и выбран договор
        self.client_id =  self.client_ids[index.row()]
        if self.hasFileFolder and self.client_id: # Сделать кнопку Сохранить активной если есть файл, папка и выбран договор
            self.clbSave.setEnabled(True)
        else:
            self.clbSave.setEnabled(False)

    def click_cbFolder(self):
        if len(self.cbFolder.currentText()):
            self.hasFileFolder = True
        else:
            self.hasFileFolder = False
        if self.hasFileFolder and self.client_id:
            self.clbSave.setEnabled(True)
        else:
            self.clbSave.setEnabled(False)

    def leFile_changed(self):
        # Соптимизировали неиспользующийся огромный словарь
        q3 = """
        self.hasFileFolder = False
        self.clbSave.setEnabled(False)
        if self.alone_files.get(self.leFile.text(), None):
            self.cbFolder.clear()
            self.cbFolder.addItems(self.alone_files[self.leFile.text()])
        """

    def click_clbRefresh(self):
        if self.calBirtday.dateTime().toPyDateTime().date() > date(1930,1,1) and \
                                self.calBirtday.dateTime().toPyDateTime() < datetime.now():
            self.leFile.setText('')
            self.hasFileFolder = False
            self.client_id = None
            self.clbSave.setEnabled(False)
            dbconn = MySQLConnection(**self.dbconfig_crm)
            cursor = dbconn.cursor()
            sql = 'SELECT cl.p_surname,cl.p_name,cl.p_lastname,cl.p_service_address,cl.d_service_address,' \
                  'ca.client_phone,ca.call_comment,ca.inserted_date,cl.client_id FROM saturn_crm.clients AS cl ' \
                  'LEFT JOIN saturn_crm.contracts AS co ON co.client_id = cl.client_id ' \
                  'LEFT JOIN saturn_crm.callcenter AS ca ON ca.contract_id = co.id ' \
                  'WHERE cl.b_date = %s'
            cursor.execute(sql, (self.calBirtday.dateTime().toPyDateTime(),))
            rows = cursor.fetchall()
            #rows = [('МЕЛЬНИКОВ', 'ВАЛЕНТИН', 'ИВАНОВИЧ', 'Тверская обл, Калязинский р-н, г Калязин, ул 1 Мая, д 5', 'Тверская обл, Калязинский р-н, г Калязин, ул 1 Мая, д 5', 79857795218, 'нпф название не помнит', datetime(2016, 4, 4, 13, 24, 46), '64c46542-fa4e-11e5-9847-5254004b76e6'), ('ХОДЕНЁВА', 'НАТАЛЬЯ', 'СЕРГЕЕВНА', 'Алтайский край, Романовский р-н, село Сидоровка, ул Партизанская, д 11А, кв 2', 'Алтайский край, г Барнаул, ул Антона Петрова, д 176, кв 50', 79237160083, '', datetime(2016, 5, 14, 8, 31, 1), '1272d548-130a-11e6-8b81-5254004b76e6')]
            dogovors = {}
            for row in rows:
                client_id = row[8]
                if dogovors.get(client_id, None):
                    if row[7].date() not in dogovors[client_id]['Даты']:
                        dogovors[client_id]['Даты'] = dogovors[client_id]['Даты'] + [row[7].date()]
                else:
                    dogovor = {}
                    dogovor['client_id'] = client_id
                    dogovor['Фамилия'] = row[0]
                    dogovor['Имя'] = row[1]
                    dogovor['Отчество'] = row[2]
                    dogovor['Регистрация'] = row[3]
                    dogovor['Проживание'] = row[4]
                    dogovor['Телефон'] = row[5]
                    dogovor['Коментарий'] = row[6]
                    if row[7]:
                        dogovor['Даты'] = [row[7].date()]
                    else:
                        dogovor['Даты'] = [None]
                    dogovors[client_id] = dogovor
            self.contracts = {}
            contracts4order = {}
            for client_id in dogovors:
                if dogovors[client_id]['Даты'] != [None]:
                    self.contracts[client_id] = dogovors[client_id]
                    contracts4order[client_id] = dogovors[client_id]['Фамилия']
            keys = ['Фамилия', 'Имя', 'Отчество', 'Регистрация', 'Проживание', 'Телефон', 'Коментарий', 'Даты']
            self.twRez.setColumnCount(len(keys))  # Устанавливаем кол-во колонок
            self.twRez.setRowCount(len(contracts4order))  # Кол-во строк из таблицы
            contracts_ordered = OrderedDict(sorted(contracts4order.items(), key=lambda t: t[1]))
            self.client_ids = []
            for j, client_id in enumerate(contracts_ordered):
                self.client_ids.append(client_id)
                for k, key in enumerate(keys):
                    if key == 'Даты':
                        if self.contracts[client_id].get('Даты', False):
                            all_dates = ';'.join([data.strftime('%d.%m.%y') for data in self.contracts[client_id][key]])
                            self.twRez.setItem(j, k, QTableWidgetItem(all_dates))
                    else:
                        self.twRez.setItem(j, k, QTableWidgetItem(str(self.contracts[client_id][key])))
            # Устанавливаем заголовки таблицы
            self.twRez.setHorizontalHeaderLabels(list(keys))
            # Устанавливаем выравнивание на заголовки
            self.twRez.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
            # делаем ресайз колонок по содержимому
            self.twRez.horizontalHeader().resizeSection(0, 150)
            self.twRez.horizontalHeader().resizeSection(1, 100)
            self.twRez.horizontalHeader().resizeSection(2, 150)
            self.twRez.horizontalHeader().resizeSection(3, 250)
            self.twRez.horizontalHeader().resizeSection(4, 250)
            self.twRez.horizontalHeader().resizeSection(5, 100)
            self.twRez.horizontalHeader().resizeSection(6, 100)
            self.twRez.horizontalHeader().resizeSection(7, 100)
            return

    def click_clbSave(self):
        self.clbReport2xlsx.setEnabled(False)
        dbconn = MySQLConnection(**self.dbconfig_alone)
        cursor = dbconn.cursor()
        sql = 'SELECT * FROM alone_connect WHERE path = %s AND file = %s AND client_id = %s'
        cursor.execute(sql, (self.cbFolder.currentText(), self.leFile.text(), self.client_id))
        rows = cursor.fetchall()
        if len(rows) == 0:
            cursor = dbconn.cursor()
            cursor.execute('INSERT INTO alone_connect (path, file, client_id) VALUES(%s, %s, %s)',
                           (self.cbFolder.currentText(), self.leFile.text(), self.client_id))
            dbconn.commit()
        dbconn.close()

    def click_pbSortF(self):
        contracts4order = {}
        for client_id in self.contracts:
            if self.contracts[client_id]['Даты'] != [None]:
                contracts4order[client_id] = self.contracts[client_id]['Фамилия']
        keys = ['Фамилия', 'Имя', 'Отчество', 'Регистрация', 'Проживание', 'Телефон', 'Коментарий', 'Даты']
        self.twRez.setColumnCount(len(keys))  # Устанавливаем кол-во колонок
        self.twRez.setRowCount(len(contracts4order))  # Кол-во строк из таблицы
        contracts_ordered = OrderedDict(sorted(contracts4order.items(), key=lambda t: t[1]))
        self.client_ids = []
        for j, client_id in enumerate(contracts_ordered):
            self.client_ids.append(client_id)
            for k, key in enumerate(keys):
                if key == 'Даты':
                    if self.contracts[client_id].get('Даты', False):
                        all_dates = ';'.join([data.strftime('%d.%m.%y') for data in self.contracts[client_id][key]])
                        self.twRez.setItem(j, k, QTableWidgetItem(all_dates))
                else:
                    self.twRez.setItem(j, k, QTableWidgetItem(str(self.contracts[client_id][key])))
        # Устанавливаем заголовки таблицы
        self.twRez.setHorizontalHeaderLabels(list(keys))
        # Устанавливаем выравнивание на заголовки
        self.twRez.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twRez.horizontalHeader().resizeSection(0, 150)
        self.twRez.horizontalHeader().resizeSection(1, 100)
        self.twRez.horizontalHeader().resizeSection(2, 150)
        self.twRez.horizontalHeader().resizeSection(3, 250)
        self.twRez.horizontalHeader().resizeSection(4, 250)
        self.twRez.horizontalHeader().resizeSection(5, 100)
        self.twRez.horizontalHeader().resizeSection(6, 100)
        self.twRez.horizontalHeader().resizeSection(7, 100)

    def click_pbSortO(self):
        contracts4order = {}
        contracts4orderNone = {}
        for client_id in self.contracts:
            if self.contracts[client_id]['Даты'] != [None]:
                if self.contracts[client_id]['Отчество']:
                    contracts4order[client_id] = self.contracts[client_id]['Отчество']
                else:
                    contracts4orderNone[client_id] = self.contracts[client_id]['Отчество']
        keys = ['Фамилия', 'Имя', 'Отчество', 'Регистрация', 'Проживание', 'Телефон', 'Коментарий', 'Даты']
        self.twRez.setColumnCount(len(keys))  # Устанавливаем кол-во колонок
        self.twRez.setRowCount(len(contracts4order))  # Кол-во строк из таблицы
        contracts_ordered = OrderedDict(sorted(contracts4order.items(), key=lambda t: t[1]))
        for client_id in contracts4orderNone:
            contracts_ordered[client_id] = contracts4orderNone[client_id]
        self.client_ids = []
        for j, client_id in enumerate(contracts_ordered):
            self.client_ids.append(client_id)
            for k, key in enumerate(keys):
                if key == 'Даты':
                    if self.contracts[client_id].get('Даты', False):
                        all_dates = ';'.join([data.strftime('%d.%m.%y') for data in self.contracts[client_id][key]])
                        self.twRez.setItem(j, k, QTableWidgetItem(all_dates))
                else:
                    self.twRez.setItem(j, k, QTableWidgetItem(str(self.contracts[client_id][key])))
        # Устанавливаем заголовки таблицы
        self.twRez.setHorizontalHeaderLabels(list(keys))
        # Устанавливаем выравнивание на заголовки
        self.twRez.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twRez.horizontalHeader().resizeSection(0, 150)
        self.twRez.horizontalHeader().resizeSection(1, 100)
        self.twRez.horizontalHeader().resizeSection(2, 150)
        self.twRez.horizontalHeader().resizeSection(3, 250)
        self.twRez.horizontalHeader().resizeSection(4, 250)
        self.twRez.horizontalHeader().resizeSection(5, 100)
        self.twRez.horizontalHeader().resizeSection(6, 100)
        self.twRez.horizontalHeader().resizeSection(7, 100)

    def click_pbSortIO(self):
        contracts4order = {}
        for client_id in self.contracts:
            if self.contracts[client_id]['Даты'] != [None]:
                contracts4order[client_id] = self.contracts[client_id]['Имя']
        keys = ['Фамилия', 'Имя', 'Отчество', 'Регистрация', 'Проживание', 'Телефон', 'Коментарий', 'Даты']
        self.twRez.setColumnCount(len(keys))  # Устанавливаем кол-во колонок
        self.twRez.setRowCount(len(contracts4order))  # Кол-во строк из таблицы
        contracts_ordered = OrderedDict(sorted(contracts4order.items(), key=lambda t: t[1]))
        self.client_ids = []
        for j, client_id in enumerate(contracts_ordered):
            self.client_ids.append(client_id)
            for k, key in enumerate(keys):
                if key == 'Даты':
                    if self.contracts[client_id].get('Даты', False):
                        all_dates = ';'.join([data.strftime('%d.%m.%y') for data in self.contracts[client_id][key]])
                        self.twRez.setItem(j, k, QTableWidgetItem(all_dates))
                else:
                    self.twRez.setItem(j, k, QTableWidgetItem(str(self.contracts[client_id][key])))
        # Устанавливаем заголовки таблицы
        self.twRez.setHorizontalHeaderLabels(list(keys))
        # Устанавливаем выравнивание на заголовки
        self.twRez.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twRez.horizontalHeader().resizeSection(0, 150)
        self.twRez.horizontalHeader().resizeSection(1, 100)
        self.twRez.horizontalHeader().resizeSection(2, 150)
        self.twRez.horizontalHeader().resizeSection(3, 250)
        self.twRez.horizontalHeader().resizeSection(4, 250)
        self.twRez.horizontalHeader().resizeSection(5, 100)
        self.twRez.horizontalHeader().resizeSection(6, 100)
        self.twRez.horizontalHeader().resizeSection(7, 100)

    def click_clbRefreshReport(self):
        q1 = """
        # Заполняем информацию по прослушиванию оператором
        dbconn = MySQLConnection(**self.dbconfig_alone)
        cursor = dbconn.cursor()
        cursor.execute('SELECT client_id, path, file FROM alone_connect')
        rows = cursor.fetchall()
        #rows = [('39f07f6d-16e7-11e8-86b5-5254004b76e6', '1', 'f658161664'), ('8113962c-16b8-11e8-86b5-5254004b76e6', '2', 'f659636224'), ('42610b96-16ea-11e8-86b5-5254004b76e6', '2', 'f659652608'), ('f51b5fd6-178d-11e8-86b5-5254004b76e6', '2', 'f687063040'), ('8c95a423-bbc5-11e6-b8cb-20cf300dec24', '6', 'f1083621376'), ('820c39ab-178f-11e8-86b5-5254004b76e6', '6', 'f1083621376'), ('3c0652e8-1809-11e8-81ec-5254004b76e6', '7', 'f2880913408'), ('d2d14811-18a0-11e8-81ec-5254004b76e6', '8', 'f3592290304'), ('b77d6b73-04ae-11e7-9f62-5254004b76e6', '9', 'f3592732672'), ('d30b3605-180c-11e8-81ec-5254004b76e6', '9', 'f3592732672'), ('525e4b86-d737-11e6-aa92-20cf300dec24', '9', 'f3712696320'), ('d8fa8330-178f-11e8-86b5-5254004b76e6', '9', 'f3712696320'), ('d3950684-fd16-11e8-8408-000c290cfc84', '25', 'f3909402624'), ('2c80db53-fddc-11e8-8408-000c290cfc84', '25', 'f3899719680'), ('c3b71aed-1da7-11e7-8786-5254004b76e6', '150', 'f2267316224'), ('f2c3a29d-156f-11e8-9039-5254004b76e6', '150', 'f2267807744'), ('d758b226-1808-11e8-81ec-5254004b76e6', '150', 'f2267807744'), ('3830045c-047b-11e9-a9ee-000c290cfc84', '150', 'f2267807744'), ('e4eba1e7-9010-11e7-8989-5254004b76e6', '3', 'f920420352'), ('f8262d02-ec50-11e7-897e-5254004b76e6', '3', 'f920420352')]
        temp_ids = []
        report_client_ids = {} # Даже индекс из папки+файл тоже может повторяться ((( Добавляем номер дубля (i) вначале
        for row in rows:
            temp_ids.append(row[0])
            for i in range (0,9):
                if report_client_ids.get(str(i) + '{0:04d}'.format(int(row[1])) + row[2], None):
                    pass
                else:
                    report_client_ids[str(i) + '{0:04d}'.format(int(row[1])) + row[2]] = row[0]
                    break
        uniq_client_ids = list(set(temp_ids)) # Убираем повторы из массива idшников чтобы
                                              # запросить внутренности нужных договоров
        dbconn = MySQLConnection(**self.dbconfig_crm)
        cursor = dbconn.cursor()
        sql = 'SELECT cl.p_surname,cl.p_name,cl.p_lastname,cl.p_service_address,cl.d_service_address,' \
              'ca.client_phone,ca.call_comment,ca.inserted_date,cl.client_id,cl.b_date FROM saturn_crm.clients AS cl ' \
              'LEFT JOIN saturn_crm.contracts AS co ON co.client_id = cl.client_id ' \
              'LEFT JOIN saturn_crm.callcenter AS ca ON ca.contract_id = co.id ' \
              'WHERE cl.client_id in ({c})'.format(c=', '.join(['%s'] * len(uniq_client_ids)))
        cursor.execute(sql, tuple(uniq_client_ids))
        rows = cursor.fetchall()
        #rows = [('ЮРЬЕВ', 'МИХАИЛ', 'АНАТОЛЬЕВИЧ', 'Омская обл, г Омск, Советский округ, ул Химиков, д 16, кв 61', 'Омская обл, г Омск, Советский округ, ул Химиков, д 16, кв 61', 79831198154, '', datetime(2018, 12, 12, 12, 54, 2), '2c80db53-fddc-11e8-8408-000c290cfc84', date(1989, 6, 15))]
        dogovors = {}
        for row in rows:
            client_id = row[8]
            if dogovors.get(client_id, None):
                if row[7].date() not in dogovors[client_id]['Даты']:
                    dogovors[client_id]['Даты'] = dogovors[client_id]['Даты'] + [row[7].date()]
            else:
                dogovor = {}
                dogovor['client_id'] = client_id
                dogovor['Фамилия'] = row[0]
                dogovor['Имя'] = row[1]
                dogovor['Отчество'] = row[2]
                dogovor['Регистрация'] = row[3]
                dogovor['Проживание'] = row[4]
                dogovor['Телефон'] = row[5]
                dogovor['Коментарий'] = row[6]
                if row[7]:
                    dogovor['Даты'] = [row[7].date()]
                else:
                    dogovor['Даты'] = [None]
                dogovor['ДеньРождения'] = row[9]
                dogovors[client_id] = dogovor
        report = {}
        for report_client_id in report_client_ids:
            path = int(report_client_id[1:5])  #file = report_client_id[5:]
            client_id = report_client_ids[report_client_id]
            dates = dogovors[client_id]['Даты']
            if report.get(path, None):
                # есть такая папка
                if report[path].get(client_id, None):
                    # есть такая папка и такой client_id
                    for data in report[path][client_id]:
                        if data not in dates:
                            dates = dates + [data]
                    report[path][client_id] = dates
                else:
                    # есть такая папка и нет такого client_id !!! первая дата - телефон
                    report[path][client_id] = [dogovors[client_id]['Телефон']] + dates
            else:
                # нет такой папки !!! первая дата - телефон
                report[path] = {client_id: [dogovors[client_id]['Телефон']] + dates}
        # перестраиваем с client_id на телефоны
        report2phones = {}
        for path in report:
            report2phones[path] = {}
            for client_id in report[path]:
                phone = report[path][client_id][0]
                if report2phones[path].get(phone, None):
                    dates = report2phones[path][phone]
                    # есть такая папка и такой телефон
                    for i, data in enumerate(report[path][client_id]):
                        if i:
                            if data not in dates:
                                dates = dates + [data]
                    report2phones[path][phone] = dates
                else:
                    # есть такая папка и нет такого телефона
                    report2phones[path][phone] = report[path][client_id][1:]
        # анализируем отчет
        self.report_rez = {}
        for path in report2phones:
            dates = {}
            for phone in report2phones[path]:
                for data in report2phones[path][phone]:
                    if dates.get(data, None):
                        # есть такая дата
                        dates[data] += 1
                    else:
                        dates[data] = 1
            dates_ordered = OrderedDict(sorted(dates.items(), key=lambda t: t[1], reverse=True))
            for data in dates_ordered:
                if len(report2phones[path]) > 1 and dates_ordered[data] >= len(report2phones[path]):
                    self.report_rez[path] = datetime.combine(data,time(0,0,0,0)).strftime('%d.%m.%y')
                elif len(report2phones[path]) > 1:
                    self.report_rez[path] = 'МУЛЬТИ'
                else:
                    self.report_rez[path] = 'начато'
                break
        """
        # Добавляем/заменяем информацию по распознанным файлам
        dbconn = MySQLConnection(**self.dbconfig_crm)
        cursor = dbconn.cursor()
        # Распознанные файлы, отсортированные по директориям
        cursor.execute('SELECT r.`path`, c.inserted_date FROM lekarh.alone_remont AS r '
                       'LEFT JOIN saturn_crm.callcenter AS c ON r.callcenter_id = c.id ORDER BY r.`path`')
        rows = cursor.fetchall()
        path = int(rows[0][0])
        count = 1
        call_dates = [rows[0][1].date()]
        self.report_rez = {}
        for i, row in enumerate(rows):
            monodate = ''
            if i:
                if int(row[0]) == path:
                    # Папка не поменялась
                    count += 1
                    if row[1].date() not in call_dates:
                        call_dates.append(row[1].date())
                else:
                    # Следующая папка
                    if count == 0:
                        monodate = ''
                    elif count == 1:
                        monodate = 'начато'
                    elif len(call_dates) == 1:
                        monodate = call_dates[0].strftime('%d.%m.%y')
                    elif len(call_dates) == 2:
                        if abs(datetime.combine(call_dates[0],time(0,0,0,0)) - \
                                                    datetime.combine(call_dates[1],time(0,0,0,0))) < timedelta(days=2):
                            monodate = datetime.combine(call_dates[0],time(0,0,0,0)).strftime('%d-') + \
                                                   datetime.combine(call_dates[1],time(0,0,0,0)).strftime('%d.%m.%y')
                        else:
                            monodate = 'МУЛЬТИ'
                    else:
                        monodate = 'МУЛЬТИ'
                    self.report_rez[path] = monodate
                    call_dates = [row[1].date()]
                    path = int(row[0])
                    count = 1
        if count == 0:
            monodate = ''
        elif count == 1:
            monodate = 'начато'
        elif len(call_dates) == 1:
            monodate = call_dates[0].strftime('%d.%m.%y')
        elif len(call_dates) == 2:
            if datetime.combine(call_dates[0], time(0, 0, 0, 0)) - \
                    datetime.combine(call_dates[1], time(0, 0, 0, 0)) < timedelta(days=2):
                monodate = datetime.combine(call_dates[0], time(0, 0, 0, 0)).strftime('%d-') + \
                           datetime.combine(call_dates[1], time(0, 0, 0, 0)).strftime('%d.%m.%y')
            else:
                monodate = 'МУЛЬТИ'
        else:
            monodate = 'МУЛЬТИ'
        self.report_rez[path] = monodate
        # Строим нити
        dbconn = MySQLConnection(**self.dbconfig_crm)
        cursor = dbconn.cursor()
        # Распознанные файлы, отсортированные по дате
        cursor.execute('SELECT r.`path`, c.inserted_date FROM lekarh.alone_remont AS r '
                       'LEFT JOIN saturn_crm.callcenter AS c ON r.callcenter_id = c.id '
                       'ORDER BY c.inserted_date, r.`path`')
        rows = cursor.fetchall()
        path = int(rows[0][0])
        self.threads = []
        for i, row in enumerate(rows):
            if i:
                if int(row[0]) == path:
                    # Папка не поменялась
                    if not len(self.threads):
                        # Если вообще ни одной, то создаем первую нить
                        self.threads.append({'start': row[1], 'end': row[1], 'maxPath': int(row[0]),
                                        'pathsDates': {int(row[0]): {row[1]: int(row[0])}}})
                    else:
                        threadUpdated = False
                        for j, thread in enumerate(self.threads):
                            if row[1].date() != thread['end'].date() or int(row[0]) != thread['maxPath']:
                                # Дата и последняя папка не совпадает?
                                if row[1] > thread['end'] and (thread['end'] + timedelta(days=30)) > row[1] and \
                                        int(row[0]) >= thread['maxPath']:
                                    # Меньше 15 дней и директория та же или увеличилась? Добавляем в рамках этой нити
                                    self.threads[j]['end'] = row[1]
                                    self.threads[j]['maxPath'] = int(row[0])
                                    if self.threads[j]['pathsDates'].get(int(row[0]), None):
                                        self.threads[j]['pathsDates'][int(row[0])][row[1]] = int(row[0])
                                    else:
                                        self.threads[j]['pathsDates'][int(row[0])] = {row[1]: int(row[0])}
                                    threadUpdated = True
                            else:
                                threadUpdated = True
                                break
                        if not threadUpdated:
                            # Создаем новую нить
                            self.threads.append({'start': row[1], 'end': row[1], 'maxPath': int(row[0]),
                                            'pathsDates': {int(row[0]): {row[1]: int(row[0])}}})
                else:
                    # Следующая папка
                    if not len(self.threads):
                        # Если вообще ни одной, то создаем первую нить
                        self.threads.append({'start': row[1], 'end': row[1], 'maxPath': int(row[0]),
                                        'pathsDates': {int(row[0]): {row[1]: int(row[0])}}})
                    else:
                        threadUpdated = False
                        for j, thread in enumerate(self.threads):
                            if row[1].date() != thread['end'].date():
                                # Дата и последняя папка не совпадает?
                                if row[1] > thread['end'] and (thread['end'] + timedelta(days=30)) > row[1] and \
                                        int(row[0]) >= thread['maxPath']:
                                    # Меньше 15 дней и директория та же или увеличилась? Добавляем в рамках этой нити
                                    self.threads[j]['end'] = row[1]
                                    self.threads[j]['maxPath'] = int(row[0])
                                    if self.threads[j]['pathsDates'].get(int(row[0]), None):
                                        self.threads[j]['pathsDates'][int(row[0])][row[1]] = int(row[0])
                                    else:
                                        self.threads[j]['pathsDates'][int(row[0])] = {row[1]: int(row[0])}
                                    threadUpdated = True
                            else:
                                threadUpdated = True
                                break
                        if not threadUpdated:
                            # Создаем новую нить
                            # Создаем новую нить
                            self.threads.append({'start': row[1], 'end': row[1], 'maxPath': int(row[0]),
                                            'pathsDates': {int(row[0]): {row[1]: int(row[0])}}})

        keys = []
        for i in range(0, 10):
            keys.append(str(i))
        hkeys = []
        for i in range(0, 546):
            hkeys.append(str(i))
        self.twRez.setColumnCount(len(keys))  # Устанавливаем кол-во колонок
        self.twRez.setRowCount(546)  # Кол-во строк из таблицы
        for j in range(0, 546):
            for k in range(0, 10):
                if self.report_rez.get(j * 10 + k, None):
                    self.twRez.setItem(j, k, QTableWidgetItem(self.report_rez[j * 10 + k]))
                else:
                    self.twRez.setItem(j, k, QTableWidgetItem('нетинф'))
        # Устанавливаем заголовки таблицы
        self.twRez.setHorizontalHeaderLabels(keys)
        # Устанавливаем заголовки таблицы
        self.twRez.setVerticalHeaderLabels(hkeys)
        # Устанавливаем выравнивание на заголовки
        self.twRez.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twRez.resizeColumnsToContents()
        self.clbReport2xlsx.setEnabled(True)


    def click_clbRefreshReportOld(self):
        dbconn = MySQLConnection(**self.dbconfig_alone)
        cursor = dbconn.cursor()
        cursor.execute('SELECT client_id, path, file FROM alone_connect')
        rows = cursor.fetchall()
        #rows = [('39f07f6d-16e7-11e8-86b5-5254004b76e6', '1', 'f658161664'), ('8113962c-16b8-11e8-86b5-5254004b76e6', '2', 'f659636224'), ('42610b96-16ea-11e8-86b5-5254004b76e6', '2', 'f659652608'), ('f51b5fd6-178d-11e8-86b5-5254004b76e6', '2', 'f687063040'), ('8c95a423-bbc5-11e6-b8cb-20cf300dec24', '6', 'f1083621376'), ('820c39ab-178f-11e8-86b5-5254004b76e6', '6', 'f1083621376'), ('3c0652e8-1809-11e8-81ec-5254004b76e6', '7', 'f2880913408'), ('d2d14811-18a0-11e8-81ec-5254004b76e6', '8', 'f3592290304'), ('b77d6b73-04ae-11e7-9f62-5254004b76e6', '9', 'f3592732672'), ('d30b3605-180c-11e8-81ec-5254004b76e6', '9', 'f3592732672'), ('525e4b86-d737-11e6-aa92-20cf300dec24', '9', 'f3712696320'), ('d8fa8330-178f-11e8-86b5-5254004b76e6', '9', 'f3712696320'), ('d3950684-fd16-11e8-8408-000c290cfc84', '25', 'f3909402624'), ('2c80db53-fddc-11e8-8408-000c290cfc84', '25', 'f3899719680'), ('c3b71aed-1da7-11e7-8786-5254004b76e6', '150', 'f2267316224'), ('f2c3a29d-156f-11e8-9039-5254004b76e6', '150', 'f2267807744'), ('d758b226-1808-11e8-81ec-5254004b76e6', '150', 'f2267807744'), ('3830045c-047b-11e9-a9ee-000c290cfc84', '150', 'f2267807744'), ('e4eba1e7-9010-11e7-8989-5254004b76e6', '3', 'f920420352'), ('f8262d02-ec50-11e7-897e-5254004b76e6', '3', 'f920420352')]
        temp_ids = []
        report_client_ids = {} # Даже индекс из папки+файл тоже может повторяться ((( Добавляем номер дубля (i) вначале
        for row in rows:
            temp_ids.append(row[0])
            for i in range (0,9):
                if report_client_ids.get(str(i) + '{0:04d}'.format(int(row[1])) + row[2], None):
                    pass
                else:
                    report_client_ids[str(i) + '{0:04d}'.format(int(row[1])) + row[2]] = row[0]
                    break
        uniq_client_ids = list(set(temp_ids)) # Убираем повторы из массива idшников чтобы
                                              # запросить внутренности нужных договоров
        dbconn = MySQLConnection(**self.dbconfig_crm)
        cursor = dbconn.cursor()
        sql = 'SELECT cl.p_surname,cl.p_name,cl.p_lastname,cl.p_service_address,cl.d_service_address,' \
              'ca.client_phone,ca.call_comment,ca.inserted_date,cl.client_id,cl.b_date FROM saturn_crm.clients AS cl ' \
              'LEFT JOIN saturn_crm.contracts AS co ON co.client_id = cl.client_id ' \
              'LEFT JOIN saturn_crm.callcenter AS ca ON ca.contract_id = co.id ' \
              'WHERE cl.client_id in ({c})'.format(c=', '.join(['%s'] * len(uniq_client_ids)))
        cursor.execute(sql, tuple(uniq_client_ids))
        rows = cursor.fetchall()
        #rows = [('ЮРЬЕВ', 'МИХАИЛ', 'АНАТОЛЬЕВИЧ', 'Омская обл, г Омск, Советский округ, ул Химиков, д 16, кв 61', 'Омская обл, г Омск, Советский округ, ул Химиков, д 16, кв 61', 79831198154, '', datetime(2018, 12, 12, 12, 54, 2), '2c80db53-fddc-11e8-8408-000c290cfc84', date(1989, 6, 15))]
        dogovors = {}
        for row in rows:
            client_id = row[8]
            if dogovors.get(client_id, None):
                if row[7].date() not in dogovors[client_id]['Даты']:
                    dogovors[client_id]['Даты'] = dogovors[client_id]['Даты'] + [row[7].date()]
            else:
                dogovor = {}
                dogovor['client_id'] = client_id
                dogovor['Фамилия'] = row[0]
                dogovor['Имя'] = row[1]
                dogovor['Отчество'] = row[2]
                dogovor['Регистрация'] = row[3]
                dogovor['Проживание'] = row[4]
                dogovor['Телефон'] = row[5]
                dogovor['Коментарий'] = row[6]
                if row[7]:
                    dogovor['Даты'] = [row[7].date()]
                else:
                    dogovor['Даты'] = [None]
                dogovor['ДеньРождения'] = row[9]
                dogovors[client_id] = dogovor
        report = {}
        for report_client_id in report_client_ids:
            path = int(report_client_id[1:5])  #file = report_client_id[5:]
            client_id = report_client_ids[report_client_id]
            dates = dogovors[client_id]['Даты']
            if report.get(path, None):
                # есть такая папка
                if report[path].get(client_id, None):
                    # есть такая папка и такой client_id
                    for data in report[path][client_id]:
                        if data not in dates:
                            dates = dates + [data]
                    report[path][client_id] = dates
                else:
                    # есть такая папка и нет такого client_id !!! первая дата - телефон
                    report[path][client_id] = [dogovors[client_id]['Телефон']] + dates
            else:
                # нет такой папки !!! первая дата - телефон
                report[path] = {client_id: [dogovors[client_id]['Телефон']] + dates}
        # перестраиваем с client_id на телефоны
        report2phones = {}
        for path in report:
            report2phones[path] = {}
            for client_id in report[path]:
                phone = report[path][client_id][0]
                if report2phones[path].get(phone, None):
                    dates = report2phones[path][phone]
                    # есть такая папка и такой телефон
                    for i, data in enumerate(report[path][client_id]):
                        if i:
                            if data not in dates:
                                dates = dates + [data]
                    report2phones[path][phone] = dates
                else:
                    # есть такая папка и нет такого телефона
                    report2phones[path][phone] = report[path][client_id][1:]
        # анализируем отчет
        self.report_rez = {}
        for path in report2phones:
            dates = {}
            for phone in report2phones[path]:
                for data in report2phones[path][phone]:
                    if dates.get(data, None):
                        # есть такая дата
                        dates[data] += 1
                    else:
                        dates[data] = 1
            dates_ordered = OrderedDict(sorted(dates.items(), key=lambda t: t[1], reverse=True))
            for data in dates_ordered:
                if len(report2phones[path]) > 1 and dates_ordered[data] >= len(report2phones[path]):
                    self.report_rez[path] = datetime.combine(data,time(0,0,0,0)).strftime('%d.%m.%y')
                elif len(report2phones[path]) > 1:
                    self.report_rez[path] = 'МУЛЬТИ'
                else:
                    self.report_rez[path] = 'начато'
                break
        keys = []
        for i in range(0, 10):
            keys.append(str(i))
        hkeys = []
        for i in range(0, 546):
            hkeys.append(str(i))
        self.twRez.setColumnCount(len(keys))  # Устанавливаем кол-во колонок
        self.twRez.setRowCount(546)  # Кол-во строк из таблицы
        for j in range(0, 546):
            for k in range(0, 10):
                if self.report_rez.get(j * 10 + k, None):
                    self.twRez.setItem(j, k, QTableWidgetItem(self.report_rez[j * 10 + k]))
                else:
                    self.twRez.setItem(j, k, QTableWidgetItem('нетинф'))
        # Устанавливаем заголовки таблицы
        self.twRez.setHorizontalHeaderLabels(keys)
        # Устанавливаем заголовки таблицы
        self.twRez.setVerticalHeaderLabels(hkeys)
        # Устанавливаем выравнивание на заголовки
        self.twRez.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twRez.resizeColumnsToContents()
        self.clbReport2xlsx.setEnabled(True)

    def click_clbReport2xlsx(self):
        wb_log = openpyxl.Workbook(write_only=True)
        ws_log = wb_log.create_sheet('Отчет')
        keys = []
        for i in range(-1, 10):
            keys.append(str(i))
        ws_log.append(keys)
        for i in range(0, 546):
            xlsx_str = []
            xlsx_str.append(str(i))
            for j in range(0, 10):
                if self.report_rez.get(i * 10 + j, None):
                    xlsx_str.append(self.report_rez[i * 10 + j])
                else:
                    xlsx_str.append('нетинф')
            ws_log.append(xlsx_str)
        wb_log.save('Отчет.xlsx')
        return

